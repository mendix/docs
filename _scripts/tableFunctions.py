import openpyxl

# Populates a blank excel table with aliases and mapped URLs from .md files
# Should only be run once, after table exists is not needed
def populateExcelFromList(list, table):
    # Open excel file
    #workbook = openpyxl.load_workbook(filename=table)
    workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Go through list
    for item in list:
        # Sets special row number for first list item
        # TO DO - see if there's a better way to set this
        if list.index(item) == 0:
            sheet.cell(row=1, column=1, value="Title")
            sheet.cell(row=1, column=2, value="URL")
            sheet.cell(row=1, column=3, value="Front matter")
            sheet.cell(row=1, column=4, value="Aliases")
            sheet.cell(row=1, column=5, value="Product")
            sheet.cell(row=1, column=6, value="First version")
            sheet.cell(row=1, column=7, value="Last version")
            sheet.cell(row=1, column=8, value="Location in product")
            sheet.cell(row=1, column=9, value="Reported to team")
            sheet.cell(row=1, column=10, value="Follow up")
            sheet.cell(row=1, column=11, value="Notes")
            rowNumber = 2
        # For every subsequent list item calculate the row number
        else:
            rowNumber = rowNumber + 1
        # Put in values from list to cells
        scratchedCell = openpyxl.styles.PatternFill("lightGrid",fill_type=None,fgColor="00C0C0C0")
        sheet.cell(row=rowNumber, column=1, value=item["Title"])
        sheet.cell(row=rowNumber, column=2, value=item["URL"])
        sheet.cell(row=rowNumber, column=3, value=item["Front matter"])
        sheet.cell(row=rowNumber, column=4, value="").fill = scratchedCell
        if item["aliases"] != None:
            # Aliases jump to a new blank row
            # There can be more than one
            for alias in item["aliases"]:
                rowNumber = rowNumber + 1
                sheet.cell(row=rowNumber, column=1, value="").fill = scratchedCell
                sheet.cell(row=rowNumber, column=2, value="").fill = scratchedCell
                sheet.cell(row=rowNumber, column=3, value="").fill = scratchedCell
                sheet.cell(row=rowNumber, column=4, value=alias)
    # Save the excel file
    workbook.save(filename=table)

# Creates a list for comparison from excel table
# The list matches the format of initial .md file parsing
# which enables comparison of this and the initial list
def createListFromExcel(table):
    # Open excel file
    workbook = openpyxl.load_workbook(filename=table)

    sheet = workbook.active
    startRow = 2
    listFromTable = []

    # Loop to keep going through rows until the last one is reached
    while startRow < sheet.max_row:

        aliases = []
        nextRow = startRow + 1
        # Grab values from cells
        title = sheet.cell(row=startRow, column=1).value
        url = sheet.cell(row=startRow, column=2).value
        mapped = sheet.cell(row=startRow, column=3).value
        if mapped == None:
            mapped = ""
        # Aliases jump to a new blank row
        # There can be more than one
        checkNext = sheet.cell(row=nextRow, column=1).value
        alias = sheet.cell(row=nextRow, column=4).value
        while (checkNext == None) and (alias != None):
            aliases.append(alias)
            nextRow = nextRow + 1
            alias = sheet.cell(row=nextRow, column=4).value
            checkNext = sheet.cell(row=nextRow, column=1).value
        # Dump all grabbed values into dict for list
        itemDict = {"Title": title, "URL": url, "Front matter": mapped, "aliases": aliases}
        listFromTable.append(itemDict)
        startRow = nextRow
    # Save the excel file
    workbook.save(filename=table)
    # Return the new list
    return listFromTable
